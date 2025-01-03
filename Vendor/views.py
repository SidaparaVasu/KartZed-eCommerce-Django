import os, io, random, string, imghdr, csv, chardet
import pandas as pd
from django.conf import settings
from PIL import Image
from django.shortcuts import get_object_or_404, render, redirect, HttpResponse
from django.urls import reverse
from django.shortcuts import render
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from .charts import *
import openpyxl
from .models import Games, GameImages, Vendor_Contact
from .forms import GameImageForm
from Authapp.models import Vendors
from Administrator.models import *
from Administrator.views import get_os_by_category

# Create your views here.

#call main page
def index_vendor(request):
    if request.session.get('is_vendor_authenticated', False):
        g_name_list,g_availstock = pie_chart(request)
        hist_price = price_histogram(request)

        return render(request,'index-vendor.html', 
        context = {
            'g_name_list' : g_name_list,'g_availstock' : g_availstock,
            'hist_price' : hist_price,
            })
    else:
        return render(request, 'vendor-login.html')    

def render_vendor_register_page(request):
    return render(request, 'vendor-register.html')

def render_vendor_login_page(request):
    return render(request, 'vendor-login.html')

def show_games_page(request):
    if request.session.get('is_vendor_authenticated', False):
        vendor = Vendors.objects.get(vendor_unique_keyid = request.session['vendor_unique_keyid'])
        games = Games.objects.filter(vendor_reference = vendor.vendor_id).order_by('game_name')
        return render(request, 'Games/show-games.html', context = {'Games':games})
    else:
        return render(request, 'vendor-login.html') 
    
def show_game_details(request, prod_key):
    if request.session.get('is_vendor_authenticated', False):
        game = Games.objects.get(product_key = prod_key)
        
        images = GameImages.objects.filter(game_id = game.gid)
        return render(request, 'Games/view-game-details.html', context = {'Game':game, 'Images':images})
    else:
        return render(request, 'vendor-login.html') 
    
def delete_game(request, product_key):
    obj = get_object_or_404(Games, product_key=product_key)
    # return HttpResponse(obj)
    if request.method == "GET":
        if obj.delete():
            messages.success(request,"Game deleted successfully!")
            return redirect(reverse(show_games_page))
        else:
            messages.error(request,"Game couldn't delete!")
    return redirect(reverse(show_games_page))

""" Check Image is in Image format or not? """
def is_image(file):
    """
    Returns True if the selected file is an image, False otherwise.
    """
    image_formats = ('.jpg', '.jpeg', '.png', '.gif')
    file_extension = os.path.splitext(file.name)[1]
    if file_extension.lower() in image_formats:
        return True
    elif imghdr.what(file) in image_formats:
        return True
    else:
        return False

# Games CRUD
def add_game_page(request):
    if request.session.get('is_vendor_authenticated', False):
        categorized_os_version_data = get_os_by_category(request)
        context = {
            'platforms'                   : Platform.objects.order_by('platform_name'),
            'game_features'               : GameFeatures.objects.order_by('game_feature_name'),
            'game_modes'                  : GameModes.objects.order_by('game_mode_name'),
            'game_categories'             : GameCategory.objects.order_by('game_category_name'),
            'operatingsys'                : OperatingSystems.objects.all(),
            # 'operatingsysversion'       : OSVersions.objects.all(),
            'categorized_version_data'    : categorized_os_version_data,
            'processors'                  : Processors.objects.all(),
            'vc'                          : VideoCards.objects.all(),
            'vcv'                         : VCVersions.objects.all(),
        }
        # return HttpResponse(context['categorized_version_data'])
        return render(request,'Games/add-game.html', context)
    else:
        return render(request, 'vendor-login.html') 

def generate_product_key(request):
    """
    Generates a random product key of the specified length.
    """
    length=16
    letters_and_digits = string.ascii_uppercase + string.ascii_lowercase + string.digits
    key = ''.join(random.choice(letters_and_digits) for i in range(length))
    return key

def check_image_format(file):
    try:
        # open the file and check if it's an image
        image = Image.open(file)
        if image.format.lower() not in ['jpeg', 'jpg', 'png', 'gif']:
            raise ValidationError('File must be in image format.')
        else:
            return image
    except IOError:
        raise ValidationError('File is not an image.')

def insert_game(request):
    if request.method == 'POST': 
        vendor_unique_id = request.POST.get('vendor_unique_keyid')
        vendor_ins = Vendors.objects.get(vendor_unique_keyid = vendor_unique_id)

        # plan = Plan.objects.get(plan_points = 100)
        plan = Plan.objects.get(points = 100)
        g_points = int(request.POST.get('game_price'))

        g_points = (g_points*plan.points)/plan.amount
        
        try:
            game = Games()
            game.product_key           = generate_product_key(request)
            game.game_logo             = request.FILES['game_logo']
            game.vendor_reference      = vendor_ins  
            game.game_name             = request.POST.get('game_name')
            game.game_description      = request.POST.get('game_description')
            game.game_developer        = request.POST.get('game_developer')
            game.game_publisher        = request.POST.get('game_publisher')
            game.game_release_date     = request.POST.get('game_release_date')
            game.avail_stock           = request.POST.get('avail_stock')
            game.game_price            = request.POST.get('game_price')
            game.discount              = request.POST.get('discount')
            game.game_points           = g_points
            game.game_storage          = request.POST.get('game_storage')
            game.game_ram              = request.POST.get('game_ram')
            game.game_features         = list(request.POST.getlist('game_features'))
            game.game_modes            = list(request.POST.getlist('game_modes'))
            game.game_categories       = list(request.POST.getlist('game_categories'))
            game.platform_names        = list(request.POST.getlist('platform_names'))
            game.game_languages        = list(request.POST.getlist('game_languages'))
            game.save()
            messages.success(request, "Game Added successfully!")
            return redirect(reverse(add_game_page))
        except Exception as e:
            messages.success(request, "Game Insertion failed!")
            return redirect(reverse(add_game_page))
    messages.error(request, "Bad request of form! Try again later!")
    return redirect(reverse(add_game_page))


""" Insert Games using CSV upload """
def games_csv_upload(request):
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name
        if file_name.endswith('.csv'):
            # For CSV files, read the file directly
            raw_data = uploaded_file.read()
            # Detect the encoding of the file
            file_encoding = chardet.detect(raw_data)['encoding']
            # Decode the file using the detected encoding
            decoded_data = raw_data.decode(file_encoding)
            # Create a StringIO object for pandas to read from
            file_stream = io.StringIO(decoded_data)
            df = pd.read_csv(file_stream)
        elif file_name.endswith('.xls') or file_name.endswith('.xlsx'):
            # For Excel files, use pandas to read the file
            df = pd.read_excel(io.BytesIO(uploaded_file.read()))
        else:
            messages.error(request, 'File is not a CSV or Excel file')
            return redirect(reverse(add_game_page))
        
        vendor_unique_id = request.POST.get('vendor_unique_keyid')
        vendor_ins = Vendors.objects.get(vendor_unique_keyid = vendor_unique_id)
        
        # Loop through each row in the DataFrame and create a new instance of YourModel
        for index, row in df.iterrows():
            # Check for any primary key or unique constraints
            try:
                # Convert the comma-separated string to a list of strings
                features = [s.strip() for s in row['game_features'].split(',')]
                modes = [s.strip() for s in row['game_modes'].split(',')]
                categories = [s.strip() for s in row['game_categories'].split(',')]
                platforms = [s.strip() for s in row['platform_names'].split(',')]
                languages = [s.strip() for s in row['game_languages'].split(',')]
                
                obj = Games.objects.create(
                    product_key           = generate_product_key(request),
                    game_logo             = "",
                    vendor_reference      = vendor_ins,
                    game_name             = row['game_name'],
                    game_description      = row['game_description'],
                    game_developer        = row['game_developer'],
                    game_publisher        = row['game_publisher'],
                    game_release_date     = row['game_release_date'],
                    avail_stock           = row['avail_stock'],
                    game_price            = row['game_price'],
                    game_points           = round((100 * int(row['game_price'])) / 125),
                    discount              = row['discount'],
                    game_storage          = row['game_storage'],
                    game_ram              = row['game_ram'],
                    game_features         = features,
                    game_modes            = modes,
                    game_categories       = categories,
                    platform_names        = platforms,
                    game_languages        = languages
                )
            except ValidationError as e:
                # Handle any validation errors
                messages.error(request, f"Error creating object: {str(e)}")
                return redirect(reverse(add_game_page))
            except IntegrityError as e:
                # Handle any duplicate primary key or unique constraint errors
                messages.error(request, f"Error creating object: {str(e)}")
                return redirect(reverse(add_game_page))
        messages.success(request, 'CSV file uploaded successfully')
        return redirect(reverse(show_games_page))
    return redirect(reverse(show_games_page))

""" upload_game_logo """
def upload_game_logo(request, prod_key):
    game = Games.objects.get(product_key = prod_key)
    url = reverse('show_game_details', args=[prod_key])
    if request.method == "POST":
        game.game_logo = request.FILES['game_logo'] 
        game.save()
        return redirect(url)
        # return render(request, 'Games/view-game-details.html', context = {'Game': [game]})
    return redirect(url)



def export_game_data(request):
        # Get all instances of the model
        queryset = Games.objects.all()

        # Define the filename and content type for the attachment
        filename = "data.csv"
        content_type = "text/csv"
        if request.GET.get("format") == "excel":
            filename = "data.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Create a response object with the file attachment headers
        response = HttpResponse(content_type=content_type)
        response["Content-Disposition"] = f"attachment; filename={filename}"

        # Generate the CSV or Excel data and write it to the response
        if request.GET.get("format") == "excel":
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for row_index, instance in enumerate(queryset, start=1):
                worksheet.cell(row=row_index, column=1, value=instance.game_logo)
                worksheet.cell(row=row_index, column=2, value=instance.game_name)
                worksheet.cell(row=row_index, column=3, value=instance.game_description)
                worksheet.cell(row=row_index, column=4, value=instance.game_developer)
                worksheet.cell(row=row_index, column=5, value=instance.game_publisher)
                worksheet.cell(row=row_index, column=6, value=instance.game_storage)
                worksheet.cell(row=row_index, column=7, value=instance.game_languages)
                worksheet.cell(row=row_index, column=8, value=instance.game_release_date)
                worksheet.cell(row=row_index, column=9, value=instance.game_price)
                worksheet.cell(row=row_index, column=10, value=instance.avail_stock)
                worksheet.cell(row=row_index, column=11, value=instance.discount)
                worksheet.cell(row=row_index, column=12, value=instance.game_points)
                worksheet.cell(row=row_index, column=13, value=instance.game_features)
                worksheet.cell(row=row_index, column=14, value=instance.game_modes)
                worksheet.cell(row=row_index, column=15, value=instance.game_categories)
                worksheet.cell(row=row_index, column=16, value=instance.platform_names)
            workbook.save(response)
        else:
            writer = csv.writer(response)
            writer.writerow([
                "game_logo", "game_name", "game_description", "game_developer", "game_publisher",
                "game_storage", "game_languages", "game_release_date", "game_price", "avail_stock",
                "avail_stock", "discount", "game_points", "game_features", "game_modes", "game_categories",
                "platform_names"
            ])
            for instance in queryset:
                writer.writerow([
                    instance.game_logo, instance.game_name, instance.game_description, instance.game_developer, instance.game_publisher,
                    instance.game_storage, instance.game_languages, instance.game_release_date, instance.game_price, instance.avail_stock,
                    instance.discount, instance.game_points, instance.game_features, instance.game_modes, instance.game_categories, instance.platform_names
                ])

        return response


def contact_game_view(request):
    if request.session.get('is_vendor_authenticated', False):
        return render(request, 'contact.html')
    else:
        return render(request, 'vendor-login.html') 

def insert_vendor_contact(request):
    #return HttpResponse("yo")
    if request.method == 'POST':
        contact_name    = request.POST.get('contact_name')
        contact_email   = request.POST.get('contact_email')
        contact_message = request.POST.get('contact_message')
        try:
            Vendor_Contact.objects.create( 
                contact_name    = contact_name,   
                contact_email   = contact_email,  
                contact_message = contact_message,
            )
            messages.success(request,"Thank you for reaching out to us! We will be in touch with you shortly.")
        except Exception as e:
            messages.success(request,"Something Went Wrong!")
            return render(request, 'contact.html')
                
    return render(request, 'contact.html')


def bulk_image_upload(request, prod_key):
    url = reverse('show_game_details', args=[prod_key])
    if request.method == "POST":
        game = Games.objects.get(product_key = prod_key)
        # return HttpResponse(game)
        
        form = GameImageForm(request.POST, request.FILES)
        if form.is_valid():
            for image in request.FILES.getlist('images'):
                GameImages.objects.create(game=game, images=image)
        return redirect(url)
    return redirect(url)

